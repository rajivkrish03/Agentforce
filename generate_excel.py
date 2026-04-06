import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Data ──────────────────────────────────────────────────────────────────────

CAPABILITIES = [
    # CASES
    dict(id='ecc',  symbol='Ec', name='Case Classification',      full_name='Einstein Case Classification',             era='Predictive', channels='Cases',
         description='Predicts picklist and checkbox field values on new cases based on past cases to reduce manual entry.',
         storyline='1. Case arrives\n2. Einstein analyzes subject/description\n3. Fields are predicted automatically',
         business_value='Faster case triaging\nImproved data quality',
         specs='Model: AutoML\nTarget: Case Fields'),
    dict(id='ecr',  symbol='Er', name='Case Routing',             full_name='Einstein Case Routing',                    era='Predictive', channels='Cases',
         description='Uses predicted field values to ensure cases reach the right agent or queue immediately.',
         storyline='1. Einstein predicts case category\n2. Omni-Channel uses prediction to find expert\n3. Case is pushed to agent',
         business_value='Reducing transfers\nFaster first response',
         specs='Logic: Omni-Channel\nTrigger: Classification'),
    dict(id='ecw',  symbol='Ew', name='Case Wrap-Up',             full_name='Einstein Case Wrap-Up',                    era='Predictive', channels='Cases',
         description='Suggests values for fields like "Resolution" or "Root Cause" when an agent is closing a case.',
         storyline='1. Agent begins closing case\n2. Einstein suggests wrap-up values\n3. Agent confirms and saves',
         business_value='Streamlined closing\nBetter trend reporting',
         specs='Trigger: Status: Closed\nFeedback: Real-time'),
    dict(id='ear',  symbol='Ar', name='Article Recommendations',  full_name='Einstein Article Recommendations',          era='Predictive', channels='Cases, Messaging, Voice, Email, Knowledge',
         description='Automatically finds and recommends the best knowledge articles to resolve cases faster.',
         storyline='1. Agent views case\n2. Einstein searches knowledge base\n3. Top articles are surfaced in sidebar',
         business_value='Agent enablement\nConsistency in support',
         specs='Tech: Vectorized Search\nAccuracy: High'),
    dict(id='ecm',  symbol='Em', name='Conversation Mining',      full_name='Einstein Conversation Mining',              era='Predictive', channels='Cases, Messaging, Voice',
         description='Analyzes case and chat transcripts to uncover common customer problems and automation opportunities.',
         storyline='1. Manager selects transcripts\n2. Einstein extracts recurring patterns\n3. New bot intents are identified',
         business_value='Data-driven bot strategy\nProcess optimization',
         specs='Insight: Unsupervised ML\nOutput: Intent Map'),
    dict(id='ews',  symbol='Ws', name='Work Summaries',           full_name='Einstein Work Summaries',                  era='Generative', channels='Cases, Messaging, Voice, Field Service',
         description='Automatically drafts case summaries, issues, and resolutions based on interaction history.',
         storyline='1. Interaction ends\n2. Einstein drafts the wrap-up summary\n3. Agent saves to record',
         business_value='Reduced after-call work\nStandardized logging',
         specs='Model: LLM Generative\nSpeed: Near-instant'),
    dict(id='esc',  symbol='Gs', name='Enhanced Summaries',       full_name='Enhanced Summaries (Case)',                 era='Generative', channels='Cases',
         description='Provides agents with a concise overview of complex, long-running cases they are inheriting.',
         storyline='1. Agent opens old case\n2. GenAI summarizes 20+ emails/notes\n3. Agent is briefed in seconds',
         business_value='Faster case transfers\nManager escalation reviews',
         specs='Volume: Multi-document\nLength: Concise'),
    dict(id='ekc',  symbol='Kc', name='Knowledge Creation',       full_name='Einstein Knowledge Creation',              era='Generative', channels='Cases, Messaging, Email, Knowledge',
         description='Drafts articles directly from resolved cases to grow the company knowledge base organically.',
         storyline='1. Case resolved\n2. Einstein drafts a "How-to"\n3. Article is reviewed for publishing',
         business_value='Closing knowledge gaps\nCrowdsourced expertise',
         specs='Workflow: Integrated Review\nSource: Case + Conversations'),
    dict(id='esre', symbol='Sr', name='Service Replies',          full_name='Einstein Service Replies',                 era='Generative', channels='Cases, Messaging, Email',
         description='Generates personalized, grounded drafts for agents to use when replying to customers.',
         storyline='1. Customer sends query\n2. Einstein drafts perfect response\n3. Agent clicks send',
         business_value='Speeding up replies\nConsistent brand voice',
         specs='Base: Knowledge Articles\nGrounding: CRM Data'),
    dict(id='esa',  symbol='Sa', name='Search Answers',           full_name='Einstein Search Answers',                  era='Generative', channels='Cases, Knowledge',
         description='Provides a direct, generated answer at the top of search results instead of just listing links.',
         storyline='1. Agent searches "Policy"\n2. Einstein reads top article\n3. Einstein generates summary answer',
         business_value='Faster information retrieval\nReduced reading time',
         specs='Tech: RAG\nTrust: Einstein Trust Layer'),
    dict(id='asa',  symbol='Aa', name='Service Assistant',        full_name='Agentforce Service Assistant',             era='Agentic',    channels='Cases, Messaging',
         description='An intelligent companion in the Service Console that plans actions and assists agents in real-time.',
         storyline='1. Agent asks "Can you help?"\n2. Assistant plans investigation\n3. Assistant suggests deep actions',
         business_value='Expert guidance for all agents\nWorkflow automation',
         specs='Engine: Atlas Reasoning\nInterface: Sidepanel'),
    dict(id='asp',  symbol='Ap', name='Service Planner',          full_name='Agentforce Service Planner',               era='Agentic',    channels='Cases',
         description='The reasoning engine that breaks down high-level service goals into granular steps.',
         storyline='1. Complex goal identified\n2. Planner maps dependencies\n3. Planner coordinates execution',
         business_value='Managing multi-step resolutions\nCoordination of CRM tasks',
         specs='Tech: Dynamic Reasoning\nOutcome: Goal-oriented'),
    # MESSAGING
    dict(id='ebots',symbol='Eb', name='Einstein Bots',            full_name='Einstein Bots',                            era='Predictive', channels='Messaging',
         description='Intelligent chatbots that automate common tasks and deflect basic queries via messaging.',
         storyline='1. Customer types "Hello"\n2. Bot identifies intent\n3. Bot executes automated flow',
         business_value='24/7 basic support\nHandling high volumes',
         specs='Engine: NLU\nTools: Flow Builder'),
    dict(id='cs',   symbol='Cs', name='Conv Summaries',           full_name='Conversation Summaries',                   era='Generative', channels='Messaging, Voice',
         description='Generates concise summaries of ongoing or completed conversations for faster context.',
         storyline='1. Chat reaches 50+ lines\n2. Einstein creates contextual snippet\n3. New agent reads it in 2 seconds',
         business_value='Improved hand-offs\nSupervisor monitoring',
         specs='Latency: Real-time\nPrecision: Detailed'),
    dict(id='cc',   symbol='Cc', name='Conv Catch-Up',            full_name='Conversation Catch-Up',                    era='Generative', channels='Messaging, Voice',
         description='Provides a "catch-up" summary specifically for agents being transferred into a live chat.',
         storyline='1. Agent 1 transfers chat\n2. Agent 2 gets "Catch-up" alert\n3. Agent 2 responds with full context',
         business_value='Eliminating repetitive questions\nSeamless transfers',
         specs='Trigger: Transfer event\nSpeed: Instant'),
    dict(id='asa_m',symbol='As', name='Service Agent',            full_name='Agentforce Service Agent',                 era='Agentic',    channels='Messaging, Email, Knowledge',
         description='Fully autonomous AI agent that manages entire customer interactions across messaging channels.',
         storyline='1. Customer asks for refund\n2. Agent reasons over policy\n3. Agent initiates refund securely',
         business_value='End-to-end automation\nScaling without hiring',
         specs='Logic: Atlas Engine\nSafety: Einstein Trust Layer'),
    dict(id='psss', symbol='Ps', name='Proactive Service',        full_name='Proactive Service for Self-Service',       era='Agentic',    channels='Messaging',
         description='Initiates self-service help proactively based on business signals (e.g., shipment delay).',
         storyline='1. Event: Item out of stock\n2. Agent triggers messaging outreach\n3. Customer resolves via Bot',
         business_value='Reducing inbound inquiries\nImproving CSAT',
         specs='Trigger: Data Cloud Signals\nOutreach: Digital'),
    # VOICE
    dict(id='cis',  symbol='Ci', name='Conv Intelligence',        full_name='Conversation Intelligence Service (Pilot)',era='Predictive', channels='Voice',
         description='Finds key moments, action items, and trends in voice call transcripts.',
         storyline='1. Call recorded & transcribed\n2. Einstein finds "Cancellation" trend\n3. Manager adjusts policy',
         business_value='Quality assurance automation\nCoaching reps',
         specs='Status: Pilot\nTech: Audio processing'),
    dict(id='av',   symbol='Av', name='Agentforce Voice',         full_name='Agentforce Voice',                         era='Agentic',    channels='Voice',
         description='Enables autonomous AI to conduct voice-based customer support with natural language understanding.',
         storyline='1. Phone rings\n2. Vocal agent answers fluently\n3. Task resolved via voice command',
         business_value='High-scale phone support\nPersonalized IVR',
         specs='Tech: VoiceBot\nNLP: Deep Learning'),
    # EMAIL
    dict(id='eet',  symbol='Et', name='Email Templates',          full_name='Einstein Email Templates',                 era='Generative', channels='Email',
         description='Helps admins generate beautiful, context-relevant email templates in seconds.',
         storyline='1. Admin needs "Renewal" template\n2. Einstein drafts HTML & text\n3. Template deployed across org',
         business_value='Faster setup\nBetter formatting',
         specs='Type: Content creation\nGrounding: Org Metadata'),
    # FIELD SERVICE
    dict(id='nba',  symbol='Nb', name='Next Best Action',         full_name='Einstein Next Best Action',                era='Predictive', channels='Field Service',
         description='Recommends optimal up-sell and cross-sell opportunities at the moment of job completion.',
         storyline='1. Technician finishes repair\n2. Einstein suggests "Battery Upgrade"\n3. Technician adds to work order',
         business_value='Increasing field revenue\nProactive maintenance',
         specs='Tech: Recommendation Engine\nUI: App & Console'),
    dict(id='erb',  symbol='Rb', name='Recommend Builder',        full_name='Einstein Recommendation Builder',          era='Predictive', channels='Field Service',
         description='Allows builds of custom recommendation models to predict parts, skills, or objects for field jobs.',
         storyline='1. Historical work orders analyzed\n2. Einstein predicts "Skill: HVAC Senior"\n3. Dispatcher assigns correctly',
         business_value='Optimizing first-time fix rates\nResource allocation',
         specs='Tool: Declarative Builder\nTarget: Any Standard Object'),
    dict(id='pwb',  symbol='Pb', name='Pre-Work Brief',           full_name='Pre-Work Brief (GenAI)',                   era='Generative', channels='Field Service',
         description='Gathers all customer history and asset info into a short, easy-to-read brief for a mobile technician.',
         storyline='1. Technician arrives on site\n2. Opens mobile phone\n3. GenAI gives 3-point summary of asset history',
         business_value='Better prepared technicians\nEnhanced customer experience',
         specs='Output: Mobile Optimized\nSource: Asset & Case History'),
    dict(id='afsa', symbol='Af', name='FS Agent',                 full_name='Agentforce Field Service Agent',           era='Agentic',    channels='Field Service',
         description='An autonomous agent that manages technician questions and administrative field tasks.',
         storyline='1. Technician asks "How many bolts?"\n2. Agent checks PDF manual\n3. Agent supplies specs instantly',
         business_value='Mobile self-sufficiency\nReduced back-office calls',
         specs='Tech: Retrieval Augmented Plan\nUI: FSL Mobile App'),
    dict(id='asa_sched',symbol='As',name='Scheduling Agent',      full_name='Agentforce Scheduling Agent',              era='Agentic',    channels='Field Service',
         description='Manages complex rescheduling and scheduling tasks autonomously for customers.',
         storyline="1. Customer texts 'Can't make it'\n2. Agent finds new slot\n3. Agent updates job in FSL",
         business_value='Appointment flexibility\nDispatcher relief',
         specs='Logic: Optimization Engine\nFlow: Conversational'),
    # KNOWLEDGE
    dict(id='eke',  symbol='Ke', name='Knowledge Edits',          full_name='Einstein Knowledge Edits',                 era='Generative', channels='Knowledge',
         description='Helps knowledge managers rewrite, shorten, or change the tone of existing articles.',
         storyline='1. Clunky article text highlighted\n2. Einstein rewrites for clarity\n3. Manager publishes professional version',
         business_value='Faster article lifecycle\nBetter readability',
         specs='Feature: Tone & Style control\nTool: Knowledge Console'),
]

# ── Colors ────────────────────────────────────────────────────────────────────

ERA_COLORS = {
    'Predictive': {'header': '0176D3', 'light': 'D6E8F8', 'dark': '0176D3'},
    'Generative': {'header': 'B8920A', 'light': 'FFF8D6', 'dark': 'FFDB3C'},
    'Agentic':    {'header': 'A0112A', 'light': 'FAD5DD', 'dark': 'E31754'},
}

HEADER_BG   = '0B0B0F'
HEADER_FONT = 'FFFFFF'
TITLE_BG    = '0176D3'
SUBHDR_BG   = '1B1B1F'
SUBHDR_FONT = 'C0C7D4'

def hex_fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def thin_border():
    s = Side(style='thin', color='DDDDDD')
    return Border(left=s, right=s, top=s, bottom=s)

# ── Workbook ──────────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()

# ── Sheet 1: Full Analytics Table ─────────────────────────────────────────────

ws = wb.active
ws.title = 'All Features'

# Title row
ws.merge_cells('A1:J1')
t = ws['A1']
t.value = 'Salesforce AI Explorer — Service Cloud Feature Analytics'
t.font = Font(name='Calibri', bold=True, size=16, color=HEADER_FONT)
t.fill = hex_fill(TITLE_BG)
t.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

# Sub-header
ws.merge_cells('A2:J2')
s = ws['A2']
s.value = 'Service Cloud · All AI Capabilities · Source of truth for feature metadata'
s.font = Font(name='Calibri', size=10, color=SUBHDR_FONT, italic=True)
s.fill = hex_fill(SUBHDR_BG)
s.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 22

# Column headers
COLS = ['#', 'Symbol', 'Feature Name', 'Full Name', 'AI Era', 'Channels',
        'Description', 'Functional Storyline', 'Business Value', 'Tech Specs']
WIDTHS = [4, 8, 22, 34, 13, 30, 46, 44, 30, 32]

for col_i, (label, width) in enumerate(zip(COLS, WIDTHS), start=1):
    cell = ws.cell(row=3, column=col_i, value=label)
    cell.font = Font(name='Calibri', bold=True, size=10, color=HEADER_FONT)
    cell.fill = hex_fill('232329')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border()
    ws.column_dimensions[get_column_letter(col_i)].width = width

ws.row_dimensions[3].height = 28

# Data rows
for row_i, cap in enumerate(CAPABILITIES, start=4):
    ec = ERA_COLORS[cap['era']]
    bg = ec['light']

    values = [
        row_i - 3,
        cap['symbol'],
        cap['name'],
        cap['full_name'],
        cap['era'],
        cap['channels'],
        cap['description'],
        cap['storyline'],
        cap['business_value'],
        cap['specs'],
    ]

    for col_i, val in enumerate(values, start=1):
        cell = ws.cell(row=row_i, column=col_i, value=val)
        cell.font = Font(name='Calibri', size=10,
                         bold=(col_i in (2, 3)),
                         color=ec['header'] if col_i == 2 else '000000')
        cell.fill = hex_fill(bg)
        cell.alignment = Alignment(vertical='top', wrap_text=True,
                                   horizontal='center' if col_i <= 2 else 'left')
        cell.border = thin_border()

    ws.row_dimensions[row_i].height = 72

# Freeze panes
ws.freeze_panes = 'A4'

# Auto-filter on header row
ws.auto_filter.ref = f'A3:J{3 + len(CAPABILITIES)}'

# ── Sheet 2: Era Summary ───────────────────────────────────────────────────────

ws2 = wb.create_sheet('Era Summary')

ws2.merge_cells('A1:D1')
t2 = ws2['A1']
t2.value = 'Feature Count by Era × Channel'
t2.font = Font(name='Calibri', bold=True, size=14, color=HEADER_FONT)
t2.fill = hex_fill(TITLE_BG)
t2.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 32

channels = ['Cases', 'Messaging', 'Voice', 'Email', 'Field Service', 'Knowledge']
eras = ['Predictive', 'Generative', 'Agentic']

# Header row
for ci, label in enumerate(['Channel', 'Predictive', 'Generative', 'Agentic', 'Total'], start=1):
    cell = ws2.cell(row=2, column=ci, value=label)
    cell.font = Font(name='Calibri', bold=True, size=10, color=HEADER_FONT)
    ec_map = {'Predictive': ERA_COLORS['Predictive']['header'],
              'Generative': ERA_COLORS['Generative']['header'],
              'Agentic':    ERA_COLORS['Agentic']['header']}
    bg_c = ec_map.get(label, '232329')
    cell.fill = hex_fill(bg_c)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border()

ws2.column_dimensions['A'].width = 18
for col in ['B', 'C', 'D', 'E']:
    ws2.column_dimensions[col].width = 14

for ri, ch in enumerate(channels, start=3):
    ch_lower = ch.lower().replace(' ', '_')
    counts = {}
    for era in eras:
        counts[era] = sum(
            1 for c in CAPABILITIES
            if c['era'] == era and ch_lower in c['channels'].lower().replace(' ', '_')
        )
    total = sum(counts.values())

    row_vals = [ch, counts['Predictive'], counts['Generative'], counts['Agentic'], total]
    for ci, val in enumerate(row_vals, start=1):
        cell = ws2.cell(row=ri, column=ci, value=val)
        cell.font = Font(name='Calibri', size=10, bold=(ci == 1 or ci == 5))
        if ci == 1:
            cell.fill = hex_fill('F2F2F2')
        elif ci == 2:
            cell.fill = hex_fill('D6E8F8')
        elif ci == 3:
            cell.fill = hex_fill('FFF8D6')
        elif ci == 4:
            cell.fill = hex_fill('FAD5DD')
        else:
            cell.fill = hex_fill('E8E8E8')
        cell.alignment = Alignment(horizontal='center' if ci > 1 else 'left', vertical='center')
        cell.border = thin_border()
    ws2.row_dimensions[ri].height = 22

# Totals row
ri = 3 + len(channels)
totals = [sum(1 for c in CAPABILITIES if c['era'] == era) for era in eras]
row_vals = ['TOTAL', totals[0], totals[1], totals[2], sum(totals)]
for ci, val in enumerate(row_vals, start=1):
    cell = ws2.cell(row=ri, column=ci, value=val)
    cell.font = Font(name='Calibri', bold=True, size=10, color=HEADER_FONT)
    cell.fill = hex_fill('232329')
    cell.alignment = Alignment(horizontal='center' if ci > 1 else 'left', vertical='center')
    cell.border = thin_border()
ws2.row_dimensions[ri].height = 24

# ── Sheet 3: Per-Channel Detail tabs ──────────────────────────────────────────

CHANNEL_ICON = {
    'cases': 'Cases', 'messaging': 'Messaging', 'voice': 'Voice',
    'email': 'Email', 'field_service': 'Field Service', 'knowledge': 'Knowledge'
}

for ch_id, ch_label in CHANNEL_ICON.items():
    ws_ch = wb.create_sheet(ch_label)

    ws_ch.merge_cells('A1:H1')
    tc = ws_ch['A1']
    tc.value = f'{ch_label} — AI Features'
    tc.font = Font(name='Calibri', bold=True, size=13, color=HEADER_FONT)
    tc.fill = hex_fill('06A59A')
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws_ch.row_dimensions[1].height = 30

    ch_caps = [c for c in CAPABILITIES
               if ch_id in c['channels'].lower().replace(' ', '_')]

    hdrs = ['Symbol', 'Feature Name', 'AI Era', 'Description',
            'Functional Storyline', 'Business Value', 'Tech Specs']
    ch_widths = [8, 26, 14, 44, 44, 30, 30]

    for ci, (h, w) in enumerate(zip(hdrs, ch_widths), start=1):
        cell = ws_ch.cell(row=2, column=ci, value=h)
        cell.font = Font(name='Calibri', bold=True, size=10, color=HEADER_FONT)
        cell.fill = hex_fill('232329')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border()
        ws_ch.column_dimensions[get_column_letter(ci)].width = w
    ws_ch.row_dimensions[2].height = 24

    for ri, cap in enumerate(ch_caps, start=3):
        ec = ERA_COLORS[cap['era']]
        vals = [cap['symbol'], cap['name'], cap['era'], cap['description'],
                cap['storyline'], cap['business_value'], cap['specs']]
        for ci, val in enumerate(vals, start=1):
            cell = ws_ch.cell(row=ri, column=ci, value=val)
            cell.font = Font(name='Calibri', size=10,
                             bold=(ci <= 2),
                             color=ec['header'] if ci == 1 else '000000')
            cell.fill = hex_fill(ec['light'])
            cell.alignment = Alignment(vertical='top', wrap_text=True,
                                       horizontal='center' if ci <= 2 else 'left')
            cell.border = thin_border()
        ws_ch.row_dimensions[ri].height = 72

    ws_ch.freeze_panes = 'A3'
    ws_ch.auto_filter.ref = f'A2:G{2 + len(ch_caps)}'

# ── Save ──────────────────────────────────────────────────────────────────────

out = '/Users/rajiv/Agentforce/Salesforce_AI_Features.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'Total features: {len(CAPABILITIES)}')
